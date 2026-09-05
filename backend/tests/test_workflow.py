import csv
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from reconcileflow import ReconciliationRunSummary, run_reconciliation_workflow
from reconcileflow.ingestion import IngestionError
from reconcileflow.reconciliation import ReconciliationConfig


SAMPLES = Path(__file__).parents[2] / "data" / "sample"


def test_csv_end_to_end_workflow_returns_complete_summary(tmp_path):
    output = tmp_path / "reports" / "reconciliation.csv"
    summary = run_reconciliation_workflow(
        bank_path=SAMPLES / "bank_transactions.csv",
        erp_path=SAMPLES / "erp_invoices.csv",
        gateway_path=SAMPLES / "gateway_settlements.csv",
        output_path=output,
    )
    assert isinstance(summary, ReconciliationRunSummary)
    assert summary.bank_transactions_loaded == 9
    assert summary.erp_invoices_loaded == 8
    assert summary.gateway_entries_loaded == 2
    assert summary.reconciliation_results == 8
    assert summary.results_requiring_review == 3  # duplicate plus two unmatched/review outcomes
    assert summary.result_counts_by_status["REQUIRES_REVIEW"] == 2
    assert summary.output_format == "CSV"
    assert summary.output_path == output
    with output.open(encoding="utf-8-sig", newline="") as handle:
        assert len(list(csv.DictReader(handle))) == 8


def test_mixed_inputs_can_produce_xlsx_output(tmp_path):
    output = tmp_path / "reconciliation.xlsx"
    summary = run_reconciliation_workflow(
        bank_path=SAMPLES / "bank_transactions.xlsx",
        erp_path=SAMPLES / "erp_invoices.csv",
        gateway_path=SAMPLES / "gateway_settlements.xlsx",
        output_path=output,
        reconciliation=ReconciliationConfig(amount_tolerance=Decimal("1.00")),
    )
    assert summary.output_format == "XLSX"
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.active.max_row == 9  # header plus eight results
    finally:
        workbook.close()


def test_gateway_input_is_optional(tmp_path):
    summary = run_reconciliation_workflow(
        bank_path=SAMPLES / "bank_transactions.csv",
        erp_path=SAMPLES / "erp_invoices.csv",
        output_path=tmp_path / "without-gateway.csv",
    )
    assert summary.gateway_entries_loaded == 0
    assert summary.reconciliation_results > 0


def test_processing_failure_does_not_create_output(tmp_path):
    output = tmp_path / "must-not-exist.csv"
    with pytest.raises(IngestionError):
        run_reconciliation_workflow(
            bank_path=tmp_path / "missing-bank.csv",
            erp_path=SAMPLES / "erp_invoices.csv",
            output_path=output,
        )
    assert not output.exists()


def test_existing_output_is_protected(tmp_path):
    output = tmp_path / "existing.csv"
    output.write_text("keep me", encoding="utf-8")
    with pytest.raises(FileExistsError, match="overwrite=True"):
        run_reconciliation_workflow(
            bank_path=SAMPLES / "bank_transactions.csv",
            erp_path=SAMPLES / "erp_invoices.csv",
            output_path=output,
        )
    assert output.read_text(encoding="utf-8") == "keep me"


def test_unsupported_output_extension_fails_before_processing(tmp_path):
    output = tmp_path / "results.json"
    with pytest.raises(ValueError, match=r"\.csv or \.xlsx"):
        run_reconciliation_workflow(
            bank_path=tmp_path / "also-missing.csv",
            erp_path=tmp_path / "missing.xlsx",
            output_path=output,
        )
    assert not output.exists()
