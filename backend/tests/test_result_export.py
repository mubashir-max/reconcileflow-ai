import csv
import json
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from reconcileflow.exporting import (
    RESULT_COLUMNS,
    ResultExportError,
    export_results_csv,
    export_results_xlsx,
    results_to_dicts,
)
from reconcileflow.models import ReconciliationStatus
from reconcileflow.reconciliation import ReconciliationResult


def _result(result_id="RESULT-0001"):
    return ReconciliationResult(
        result_id=result_id,
        status=ReconciliationStatus.MANY_TO_ONE_MATCH,
        rule="MULTIPLE_INVOICES_TO_BANK",
        bank_source_record_ids=("BANK-0004",),
        erp_invoice_ids=("INV-1004", "INV-1005"),
        expected_amount=Decimal("3675.00"),
        actual_amount=Decimal("3675.00"),
        amount_difference=Decimal("0.00"),
        currency="AED",
        explanation="Multiple referenced invoices total the bank payment.",
        requires_review=False,
    )


def test_dict_rows_are_json_compatible_and_preserve_decimal_text():
    rows = results_to_dicts([_result()])
    assert tuple(rows[0]) == RESULT_COLUMNS
    assert rows[0]["expected_amount"] == "3675.00"
    assert rows[0]["erp_invoice_ids"] == "INV-1004|INV-1005"
    assert json.loads(json.dumps(rows))[0]["requires_review"] is False


def test_csv_and_xlsx_contain_equivalent_information(tmp_path):
    csv_path = export_results_csv([_result()], tmp_path / "csv" / "results.csv")
    xlsx_path = export_results_xlsx([_result()], tmp_path / "xlsx" / "results.xlsx")
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        csv_row = next(csv.DictReader(handle))
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        values = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    xlsx_row = dict(zip(values[0], values[1], strict=True))
    for column in RESULT_COLUMNS[:-1]:
        assert xlsx_row[column] == (csv_row[column] or None)
    assert csv_row["requires_review"] == "False"
    assert xlsx_row["requires_review"] is False


@pytest.mark.parametrize("exporter,suffix", [(export_results_csv, ".csv"), (export_results_xlsx, ".xlsx")])
def test_empty_export_contains_headers(tmp_path, exporter, suffix):
    output = exporter([], tmp_path / f"empty{suffix}")
    if suffix == ".csv":
        with output.open(encoding="utf-8-sig", newline="") as handle:
            assert list(csv.reader(handle)) == [list(RESULT_COLUMNS)]
    else:
        workbook = load_workbook(output, read_only=True, data_only=True)
        try:
            assert list(workbook.active.values) == [RESULT_COLUMNS]
        finally:
            workbook.close()


@pytest.mark.parametrize("exporter,suffix", [(export_results_csv, ".csv"), (export_results_xlsx, ".xlsx")])
def test_existing_file_is_protected_unless_overwrite_is_explicit(tmp_path, exporter, suffix):
    output = tmp_path / f"results{suffix}"
    exporter([_result()], output)
    with pytest.raises(FileExistsError, match="overwrite=True"):
        exporter([_result("RESULT-0002")], output)
    exporter([_result("RESULT-0002")], output, overwrite=True)


def test_wrong_extension_is_rejected(tmp_path):
    with pytest.raises(ResultExportError, match=r"\.csv extension"):
        export_results_csv([_result()], tmp_path / "results.xlsx")


def test_non_result_values_are_rejected():
    with pytest.raises(TypeError, match="ReconciliationResult"):
        results_to_dicts([object()])
