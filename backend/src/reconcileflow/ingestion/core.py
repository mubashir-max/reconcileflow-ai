"""Configurable, traceable ingestion for heterogeneous financial exports."""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from dataclasses import dataclass, field, fields
from datetime import date, datetime, timezone, tzinfo
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any, Iterable, Mapping, TypeVar

from openpyxl import load_workbook

from reconcileflow.models import (
    BankTransaction,
    BankTransactionStatus,
    CreditDebitIndicator,
    ERPInvoice,
    GatewaySettlementEntry,
    GatewayStatus,
    GatewayTransactionType,
    InvoiceLine,
    InvoiceStatus,
    InvoiceType,
)


class IngestionError(ValueError):
    """A data error with enough source context to fix the offending record."""

    def __init__(
        self,
        message: str,
        *,
        path: str | Path,
        row: int | None = None,
        column: str | None = None,
        value: Any = None,
    ) -> None:
        self.path = Path(path)
        self.row = row
        self.column = column
        self.value = value
        location = str(self.path)
        if row is not None:
            location += f": row {row}"
        if column is not None:
            location += f", column {column!r}"
        if value not in (None, ""):
            location += f", value {value!r}"
        super().__init__(f"{location}: {message}")


@dataclass(frozen=True, slots=True)
class IngestionConfig:
    """Per-source settings; keys are source headers and values are model fields."""

    column_mapping: Mapping[str, str] = field(default_factory=dict)
    source_timezone: tzinfo = timezone.utc
    sheet_name: str | None = None


@dataclass(frozen=True, slots=True)
class _Row:
    number: int
    values: Mapping[str, Any]


def _key(value: Any) -> str:
    return " ".join(str(value).strip().replace("_", " ").replace("-", " ").split()).casefold()


def _read(path: str | Path, config: IngestionConfig) -> tuple[list[str], list[_Row]]:
    source = Path(path)
    if not source.is_file():
        raise IngestionError("file does not exist", path=source)
    suffix = source.suffix.casefold()
    if suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            data = list(csv.reader(handle))
    elif suffix == ".xlsx":
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            if config.sheet_name is not None:
                if config.sheet_name not in workbook.sheetnames:
                    raise IngestionError(f"worksheet {config.sheet_name!r} does not exist", path=source)
                sheet = workbook[config.sheet_name]
            else:
                sheet = workbook.active
            data = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    else:
        raise IngestionError("supported formats are .csv and .xlsx", path=source)

    if not data or not any(value not in (None, "") for value in data[0]):
        raise IngestionError("file has no header row", path=source)
    headers = [str(value).strip() if value is not None else "" for value in data[0]]
    if any(not header for header in headers):
        raise IngestionError("header names cannot be blank", path=source, row=1)
    normalized = [_key(header) for header in headers]
    if len(normalized) != len(set(normalized)):
        duplicate = next(name for name in normalized if normalized.count(name) > 1)
        raise IngestionError(f"duplicate header after normalization: {duplicate!r}", path=source, row=1)
    rows = [
        _Row(number=index, values=dict(zip(headers, values, strict=False)))
        for index, values in enumerate(data[1:], start=2)
        if any(value not in (None, "") for value in values)
    ]
    return headers, rows


ALIASES: dict[str, tuple[str, ...]] = {
    "transaction_id": ("transaction id", "bank transaction id", "transaction reference", "txn id"),
    "source_record_id": ("source record id", "record id", "external record id"),
    "account_id": ("account id", "bank account", "account number"),
    "booking_date": ("booking date", "posted date", "transaction date"),
    "amount": ("amount", "transaction amount"),
    "credit_debit_indicator": ("credit debit indicator", "credit/debit", "dr cr"),
    "invoice_id": ("invoice id", "invoice key"),
    "invoice_number": ("invoice number", "invoice no", "document number", "docnum"),
    "customer_id": ("customer id", "customer code", "account code"),
    "issue_date": ("issue date", "invoice date", "document date"),
    "settlement_id": ("settlement id", "payout id"),
    "gateway_payment_id": ("gateway payment id", "payment id", "charge id"),
    "merchant_account_id": ("merchant account id", "merchant id"),
    "created_at": ("created at", "transaction timestamp"),
    "settlement_date": ("settlement date", "payout date"),
    "gross_amount": ("gross amount", "gross"),
    "fee_amount": ("fee amount", "processing fee", "fee"),
    "net_amount": ("net amount", "net"),
}


def _map_rows(
    path: Path,
    headers: list[str],
    rows: list[_Row],
    canonical: set[str],
    required: set[str],
    config: IngestionConfig,
) -> list[_Row]:
    lookup = {_key(name): name for name in canonical}
    for canonical_name, aliases in ALIASES.items():
        if canonical_name in canonical:
            for alias in aliases:
                lookup[_key(alias)] = canonical_name
    explicit = {_key(source): target for source, target in config.column_mapping.items()}
    invalid = sorted(set(explicit.values()) - canonical)
    if invalid:
        raise IngestionError(f"mapping targets unknown fields: {', '.join(invalid)}", path=path)

    resolved: dict[str, str] = {}
    for header in headers:
        target = explicit.get(_key(header), lookup.get(_key(header)))
        if target:
            if target in resolved.values():
                raise IngestionError(f"multiple source columns map to {target!r}", path=path, row=1)
            resolved[header] = target
    missing = sorted(required - set(resolved.values()))
    if missing:
        raise IngestionError(f"missing required columns: {', '.join(missing)}", path=path, row=1)

    result = []
    for row in rows:
        values = {resolved[h]: value for h, value in row.values.items() if h in resolved}
        unknown = {h: value for h, value in row.values.items() if h not in resolved and value not in (None, "")}
        values["__unknown__"] = unknown
        result.append(_Row(row.number, values))
    return result


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: Any) -> str | None:
    if _blank(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _convert(value: Any, kind: Any, *, path: Path, row: int, column: str, zone: tzinfo) -> Any:
    if _blank(value):
        return None
    try:
        if kind is str:
            return _text(value)
        if kind is Decimal:
            return Decimal(str(value).replace(",", "").strip())
        if kind is int:
            decimal = Decimal(str(value).strip())
            if decimal != decimal.to_integral_value():
                raise ValueError("must be a whole number")
            return int(decimal)
        if kind is bool:
            if isinstance(value, bool):
                return value
            choices = {"true": True, "1": True, "yes": True, "y": True, "false": False, "0": False, "no": False, "n": False}
            return choices[str(value).strip().casefold()]
        if kind is date:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            return date.fromisoformat(str(value).strip())
        if kind is datetime:
            if isinstance(value, datetime):
                parsed = value
            else:
                parsed = datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
            return parsed.replace(tzinfo=zone) if parsed.tzinfo is None else parsed
        if isinstance(kind, type) and issubclass(kind, Enum):
            return kind(str(value).strip().upper().replace(" ", "_"))
    except (InvalidOperation, ValueError, TypeError, KeyError) as exc:
        raise IngestionError(f"cannot convert to {getattr(kind, '__name__', kind)} ({exc})", path=path, row=row, column=column, value=value) from exc
    return value


ENUMS = {
    "credit_debit_indicator": CreditDebitIndicator,
    "status_bank": BankTransactionStatus,
    "invoice_type": InvoiceType,
    "status_invoice": InvoiceStatus,
    "transaction_type": GatewayTransactionType,
    "status_gateway": GatewayStatus,
}
DECIMALS = {
    "amount", "account_amount", "exchange_rate", "balance_after_transaction", "gross_amount", "fee_amount",
    "tax_on_fee_amount", "refund_amount", "chargeback_amount", "adjustment_amount", "net_amount", "quantity",
    "unit_price", "line_discount_amount", "line_charge_amount", "line_net_amount", "tax_rate", "tax_amount",
    "line_gross_amount", "document_subtotal", "document_discount_amount", "document_charge_amount",
    "document_tax_amount", "document_total_amount", "prepaid_amount", "paid_amount", "amount_due",
}
DATES = {"booking_date", "value_date", "authorized_date", "settlement_date", "available_on", "issue_date", "due_date", "service_start_date", "service_end_date", "posting_date", "last_payment_date"}
DATETIMES = {"booking_datetime", "authorized_datetime", "ingested_at", "created_at", "updated_at"}
BOOLS = {"is_fee", "is_reversal"}


def _extra(values: Mapping[str, Any], *, path: Path, row: int) -> dict[str, Any]:
    result = dict(values.get("__unknown__", {}))
    raw = values.get("extra_data_json")
    if not _blank(raw):
        try:
            parsed = json.loads(str(raw))
            if not isinstance(parsed, dict):
                raise ValueError("must contain a JSON object")
            result.update(parsed)
        except (json.JSONDecodeError, ValueError) as exc:
            raise IngestionError(f"invalid extra_data_json ({exc})", path=path, row=row, column="extra_data_json", value=raw) from exc
    return result


def _kwargs(values: Mapping[str, Any], names: Iterable[str], *, path: Path, row: int, config: IngestionConfig, status_kind: Any = None) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for name in names:
        if name not in values or _blank(values[name]):
            continue
        kind: Any = str
        if name in DECIMALS:
            kind = Decimal
        elif name in DATES:
            kind = date
        elif name in DATETIMES:
            kind = datetime
        elif name in BOOLS:
            kind = bool
        elif name == "line_number":
            kind = int
        elif name == "status" and status_kind:
            kind = status_kind
        elif name in ENUMS:
            kind = ENUMS[name]
        output[name] = _convert(values[name], kind, path=path, row=row, column=name, zone=config.source_timezone)
    return output


T = TypeVar("T")


def _load_simple(path: str | Path, model: type[T], required: set[str], config: IngestionConfig, status_kind: Any) -> list[T]:
    source = Path(path)
    headers, raw = _read(source, config)
    names = {item.name for item in fields(model)} | {"extra_data_json"}
    mapped = _map_rows(source, headers, raw, names, required, config)
    output = []
    for row in mapped:
        kwargs = _kwargs(row.values, names - {"extra_data", "extra_data_json"}, path=source, row=row.number, config=config, status_kind=status_kind)
        kwargs["extra_data"] = _extra(row.values, path=source, row=row.number)
        try:
            output.append(model(**kwargs))
        except (TypeError, ValueError) as exc:
            raise IngestionError(str(exc), path=source, row=row.number) from exc
    return output


def load_bank_transactions(path: str | Path, config: IngestionConfig | None = None) -> list[BankTransaction]:
    """Load bank records from CSV/XLSX using canonical names, aliases, or explicit mappings."""
    required = {"source_system", "source_record_id", "transaction_id", "account_id", "booking_date", "amount", "currency", "credit_debit_indicator", "status", "description", "ingested_at"}
    return _load_simple(path, BankTransaction, required, config or IngestionConfig(), BankTransactionStatus)


def load_gateway_settlements(path: str | Path, config: IngestionConfig | None = None) -> list[GatewaySettlementEntry]:
    """Load gateway settlement records from CSV or XLSX."""
    required = {"source_system", "source_record_id", "settlement_id", "gateway_payment_id", "merchant_account_id", "transaction_type", "status", "created_at", "settlement_date", "gross_amount", "fee_amount", "tax_on_fee_amount", "refund_amount", "chargeback_amount", "adjustment_amount", "net_amount", "currency", "settlement_currency", "ingested_at"}
    return _load_simple(path, GatewaySettlementEntry, required, config or IngestionConfig(), GatewayStatus)


def load_erp_invoices(path: str | Path, config: IngestionConfig | None = None) -> list[ERPInvoice]:
    """Load flattened invoice-line exports and group their rows into invoices."""
    config = config or IngestionConfig()
    source = Path(path)
    invoice_names = {item.name for item in fields(ERPInvoice)} - {"lines", "extra_data"}
    line_names = {item.name for item in fields(InvoiceLine)} - {"extra_data"}
    canonical = invoice_names | line_names | {"extra_data_json"}
    required = {"source_system", "source_record_id", "invoice_id", "invoice_number", "invoice_type", "status", "issue_date", "currency", "supplier_id", "supplier_name", "customer_id", "customer_name", "document_subtotal", "document_discount_amount", "document_charge_amount", "document_tax_amount", "document_total_amount", "prepaid_amount", "paid_amount", "amount_due", "line_id", "line_number", "item_description", "quantity", "unit_price", "line_net_amount", "tax_rate", "tax_amount", "line_gross_amount", "ingested_at"}
    headers, raw = _read(source, config)
    mapped = _map_rows(source, headers, raw, canonical, required, config)
    grouped: dict[str, list[_Row]] = defaultdict(list)
    for row in mapped:
        grouped[str(row.values["invoice_id"])].append(row)

    invoices: list[ERPInvoice] = []
    for invoice_id, rows in grouped.items():
        first = rows[0]
        lines = []
        for row in rows:
            if row.values.get("invoice_id") != first.values.get("invoice_id"):
                raise IngestionError("invoice grouping mismatch", path=source, row=row.number, column="invoice_id")
            line_kwargs = _kwargs(row.values, line_names, path=source, row=row.number, config=config)
            try:
                lines.append(InvoiceLine(**line_kwargs))
            except (TypeError, ValueError) as exc:
                raise IngestionError(str(exc), path=source, row=row.number) from exc
        kwargs = _kwargs(first.values, invoice_names, path=source, row=first.number, config=config, status_kind=InvoiceStatus)
        kwargs["invoice_type"] = _convert(first.values["invoice_type"], InvoiceType, path=source, row=first.number, column="invoice_type", zone=config.source_timezone)
        kwargs["lines"] = tuple(lines)
        kwargs["extra_data"] = _extra(first.values, path=source, row=first.number)
        try:
            invoices.append(ERPInvoice(**kwargs))
        except (TypeError, ValueError) as exc:
            raise IngestionError(str(exc), path=source, row=first.number) from exc
    return invoices
